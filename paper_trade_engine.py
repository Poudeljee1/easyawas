"""
paper_trade_engine.py — Shared core for the TJL paper trading system.

Provides:
  - Market data via yfinance (daily + 1-min bars, live quote)
  - TJL condition logic (mirrors the TradingView scanner logic)
  - Paper trade placement & state persistence (paper_trading_state.json)
  - Position monitor (TP / SL only — no time-based force-close)
  - Excel trade log (paper_trades_15day.xlsx)

Timezone policy: ALL datetime operations use America/New_York explicitly
via pytz.  No system-local time or UTC arithmetic for business logic.

Position lifetime: positions carry overnight and across weekends.
TP (+3%) and SL (-1.5%) are anchored to the original entry price and
remain active until hit.  Hold time accumulates across all calendar days.
"""

import json
import math
import os
import sys
from datetime import datetime, timezone, timedelta

# ── Optional imports (graceful degradation) ──────────────────────────────────
try:
    import yfinance as yf
    YF_OK = True
except ImportError:
    YF_OK = False
    print("[WARN] yfinance not installed. Run: pip install yfinance", file=sys.stderr)

try:
    import pytz
    # All business-logic time operations use this zone explicitly.
    # Never rely on the system clock's local timezone.
    ET = pytz.timezone("America/New_York")
    PYTZ_OK = True
except ImportError:
    PYTZ_OK = False
    print("[WARN] pytz not installed. Run: pip install pytz", file=sys.stderr)
    ET = None

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False
    print("[WARN] openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)

# ── Config ────────────────────────────────────────────────────────────────────
WORK_DIR        = r"C:\Users\USER\hello"
STATE_FILE      = os.path.join(WORK_DIR, "paper_trading_state.json")
EXCEL_FILE      = os.path.join(WORK_DIR, "paper_trades_15day.xlsx")
POSITION_SIZE   = 100.0   # $ per trade
TP_PCT          = 0.030   # +3.0 % take profit
SL_PCT          = 0.015   # -1.5 % stop loss
# No EOD force-close. Positions carry overnight until TP or SL is hit.

# ── Timezone helpers ──────────────────────────────────────────────────────────

def now_et():
    """Current wall-clock time in America/New_York. Always tz-aware."""
    if not PYTZ_OK:
        raise RuntimeError("pytz required — run: pip install pytz")
    # datetime.now(ET) is the canonical way: reads system UTC clock,
    # then converts to ET (handles DST automatically via pytz).
    return datetime.now(ET)

def today_str():
    """Today's date string (YYYY-MM-DD) in ET."""
    return now_et().strftime("%Y-%m-%d")

def et_midnight(date_):
    """Return an ET-aware datetime at midnight for the given date object."""
    return ET.localize(datetime(date_.year, date_.month, date_.day, 0, 0, 0))

# ── State management ──────────────────────────────────────────────────────────

def load_state():
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {"open_positions": {}, "daily_entries": {}, "closed_trades": []}

def save_state(state):
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, indent=2)

def already_traded_today(symbol, state):
    return symbol in state["daily_entries"].get(today_str(), [])

# ── Market data ───────────────────────────────────────────────────────────────

def get_daily_bars(symbol, count=215):
    """
    Returns list of OHLCV dicts sorted oldest-first.
    Fetches ~2 years to ensure ≥210 bars available.
    """
    if not YF_OK:
        raise RuntimeError("yfinance required")
    hist = yf.Ticker(symbol).history(period="2y", interval="1d", auto_adjust=True)
    if hist.empty:
        return []
    bars = []
    for ts, row in hist.iterrows():
        bars.append({
            "time":   int(ts.timestamp()),
            "open":   float(row["Open"]),
            "high":   float(row["High"]),
            "low":    float(row["Low"]),
            "close":  float(row["Close"]),
            "volume": int(row["Volume"]),
        })
    return bars[-count:]

def get_1min_bars(symbol):
    """
    Returns list of 1-min OHLCV dicts for the past 2 days (incl. premarket),
    sorted oldest-first, times in UTC unix seconds.
    """
    if not YF_OK:
        raise RuntimeError("yfinance required")
    hist = yf.Ticker(symbol).history(
        period="2d", interval="1m", auto_adjust=True, prepost=True
    )
    if hist.empty:
        return []
    bars = []
    for ts, row in hist.iterrows():
        unix = int(ts.timestamp()) if ts.tzinfo else int(ts.replace(tzinfo=timezone.utc).timestamp())
        bars.append({
            "time":   unix,
            "open":   float(row["Open"]),
            "high":   float(row["High"]),
            "low":    float(row["Low"]),
            "close":  float(row["Close"]),
            "volume": int(row["Volume"]),
        })
    return bars

def get_live_price(symbol):
    """Returns float or None."""
    if not YF_OK:
        raise RuntimeError("yfinance required")
    try:
        price = yf.Ticker(symbol).fast_info["last_price"]
        if price and float(price) > 0:
            return round(float(price), 4)
    except Exception:
        pass
    # Fallback: last 1-min close
    try:
        hist = yf.Ticker(symbol).history(period="1d", interval="1m", auto_adjust=True)
        if not hist.empty:
            return round(float(hist["Close"].iloc[-1]), 4)
    except Exception:
        pass
    return None

# ── TJL evaluation ────────────────────────────────────────────────────────────

def evaluate_tjl(symbol, curr_px=None):
    """
    Runs the full TJL (Trend Join Long) check for one symbol.

    Returns a result dict:
      result: "pass" | "fail_daily" | "fail_intraday" | "error"
      + all supporting fields for the JSON output
    """
    out = {
        "symbol":          symbol,
        "prev_daily_high": None,
        "prev_daily_close": None,
        "sma200":          None,
        "curr_px":         curr_px,
        "pmh":             None,
        "pmh_note":        None,
        "today_hod":       None,
        "today_hod_note":  None,
        "daily_breakout":  False,
        "intraday_breakout": None,
        "result":          "error",
        "reason":          "",
    }

    # ── 1. Daily bars ──────────────────────────────────────────────────────────
    try:
        daily = get_daily_bars(symbol, 215)
    except Exception as exc:
        out["reason"] = f"daily fetch error: {exc}"
        return out

    if len(daily) < 202:
        out["result"] = "fail_daily"
        out["reason"] = f"insufficient daily history ({len(daily)} bars, need 202)"
        return out

    # bar[-2] = yesterday (completed);  bar[-1] = today (in-progress, excluded)
    prev = daily[-2]
    out["prev_daily_high"]  = round(prev["high"],  4)
    out["prev_daily_close"] = round(prev["close"], 4)

    # SMA200: last 200 completed bars = daily[-201 : -1]
    sma_slice = daily[-201:-1]
    out["sma200"] = round(sum(b["close"] for b in sma_slice) / len(sma_slice), 2)

    # ── 2. Live price ──────────────────────────────────────────────────────────
    if curr_px is None:
        try:
            curr_px = get_live_price(symbol)
        except Exception as exc:
            out["reason"] = f"quote fetch error: {exc}"
            return out
    out["curr_px"] = curr_px

    if curr_px is None:
        out["reason"] = "live price unavailable"
        return out

    # ── 3. Daily breakout ──────────────────────────────────────────────────────
    daily_ok = (curr_px > out["prev_daily_high"] and
                out["prev_daily_close"] > out["sma200"])
    out["daily_breakout"] = daily_ok

    if not daily_ok:
        out["result"] = "fail_daily"
        if curr_px <= out["prev_daily_high"]:
            out["reason"] = f"curr_px {curr_px} <= prev_daily_high {out['prev_daily_high']}"
        else:
            out["reason"] = (f"prev_daily_close {out['prev_daily_close']} "
                             f"<= sma200 {out['sma200']}")
        return out

    # ── 4. 1-min bars for PMH + today_hod ─────────────────────────────────────
    try:
        min1 = get_1min_bars(symbol)
    except Exception as exc:
        out["result"]   = "fail_intraday"
        out["pmh_note"] = f"1min fetch error: {exc}"
        out["reason"]   = out["pmh_note"]
        return out

    # Boundary timestamps — use module-level ET (America/New_York), never system time.
    net = now_et()   # tz-aware, ET
    td  = net.date()
    ts_pm_start  = int(ET.localize(datetime(td.year, td.month, td.day,  4,  0)).timestamp())
    ts_rth_start = int(ET.localize(datetime(td.year, td.month, td.day,  9, 30)).timestamp())

    pm_bars  = [b for b in min1 if ts_pm_start  <= b["time"] < ts_rth_start]
    rth_bars = [b for b in min1 if b["time"] >= ts_rth_start]

    out["pmh"] = round(max(b["high"] for b in pm_bars), 4) if pm_bars else None
    if not pm_bars:
        out["pmh_note"] = "no_premarket_bars"

    # today_hod: all completed RTH bars (exclude last in-progress bar)
    completed_rth = rth_bars[:-1] if len(rth_bars) > 1 else []
    out["today_hod"] = round(max(b["high"] for b in completed_rth), 4) if completed_rth else None
    if not completed_rth:
        out["today_hod_note"] = "no_completed_rth_bars"

    # ── 5. Intraday breakout ───────────────────────────────────────────────────
    pmh = out["pmh"]
    hod = out["today_hod"]

    if pmh is None or hod is None:
        out["intraday_breakout"] = False
        out["result"] = "fail_intraday"
        out["reason"]  = f"missing intraday data (pmh={pmh}, hod={hod})"
        return out

    intraday_ok = (curr_px > pmh and curr_px > hod)
    out["intraday_breakout"] = intraday_ok

    if intraday_ok:
        out["result"] = "pass"
        out["reason"] = "daily + intraday breakout confirmed"
    else:
        out["result"] = "fail_intraday"
        if curr_px <= pmh:
            out["reason"] = f"curr_px {curr_px} <= pmh {pmh}"
        else:
            out["reason"] = f"curr_px {curr_px} <= today_hod {hod}"

    return out

# ── Paper trade placement ─────────────────────────────────────────────────────

def place_paper_trade(symbol, entry_price, state):
    """
    Opens a simulated paper position. Mutates state (caller must save_state).
    Returns the position dict.
    """
    shares     = math.floor((POSITION_SIZE / entry_price) * 100) / 100.0  # 2dp
    tp_price   = round(entry_price * (1.0 + TP_PCT),  4)
    sl_price   = round(entry_price * (1.0 - SL_PCT),  4)
    net        = now_et()
    entry_unix = int(net.timestamp())

    position = {
        "symbol":        symbol,
        "entry_price":   entry_price,
        "shares":        shares,
        "position_size": round(shares * entry_price, 2),
        "tp_price":      tp_price,
        "sl_price":      sl_price,
        "entry_time_et": net.strftime("%Y-%m-%d %H:%M ET"),
        "entry_unix":    entry_unix,
    }

    state["open_positions"][symbol] = position

    td = today_str()
    state["daily_entries"].setdefault(td, [])
    if symbol not in state["daily_entries"][td]:
        state["daily_entries"][td].append(symbol)

    print(f"  [TRADE] BUY  {shares:.2f} {symbol} @ ${entry_price:.4f} | "
          f"TP=${tp_price:.4f}  SL=${sl_price:.4f} | "
          f"Size=${position['position_size']:.2f}")
    return position

# ── Position monitoring ───────────────────────────────────────────────────────

def check_and_close_positions(state):
    """
    Checks every open position against its live price.

    Closes ONLY on TP or SL — never on time.  Positions survive overnight,
    weekends, and holidays; they remain in paper_trading_state.json until
    the price threshold is reached on any future trading day.

    Hold time is total elapsed wall-clock minutes from entry_unix to now,
    spanning all calendar days (including non-trading hours).

    Mutates state in-place (caller must call save_state).
    Returns list of dicts for trades closed this cycle.
    """
    net    = now_et()   # ET-aware, America/New_York
    closed = []

    for symbol in list(state["open_positions"].keys()):
        pos     = state["open_positions"][symbol]
        curr_px = get_live_price(symbol)
        if curr_px is None:
            print(f"  [WARN] No price for {symbol} — skipping this cycle")
            continue

        ep = pos["entry_price"]
        tp = pos["tp_price"]
        sl = pos["sl_price"]
        sh = pos["shares"]

        # Only TP or SL triggers a close — no time-based exit
        if curr_px >= tp:
            exit_reason, result_type = "TP", "WIN"
        elif curr_px <= sl:
            exit_reason, result_type = "SL", "LOSS"
        else:
            continue  # position still open, leave it

        pnl       = round((curr_px - ep) * sh, 2)
        ret_pct   = round((curr_px - ep) / ep * 100, 2)
        # Total minutes held, accumulating across all calendar days
        hold_mins = round((net.timestamp() - pos["entry_unix"]) / 60, 1)

        trade = {
            "date":          net.strftime("%Y-%m-%d"),
            "symbol":        symbol,
            "entry_price":   ep,
            "exit_price":    round(curr_px, 4),
            "shares":        sh,
            "position_size": pos["position_size"],
            "pnl":           pnl,
            "return_pct":    ret_pct,
            "hold_mins":     hold_mins,
            "result":        result_type,
            "exit_reason":   exit_reason,
            "entry_time_et": pos["entry_time_et"],
            "exit_time_et":  net.strftime("%Y-%m-%d %H:%M ET"),
        }

        tag = "+" if result_type == "WIN" else "-"
        print(f"  [CLOSE] [{tag}] {symbol} @ ${curr_px:.4f} | {result_type} | "
              f"P&L ${pnl:+.2f} ({ret_pct:+.2f}%) | "
              f"held {hold_mins:.0f}m ({hold_mins/60:.1f}h) | {exit_reason}")

        del state["open_positions"][symbol]
        state["closed_trades"].append(trade)
        closed.append(trade)
        append_trade_to_excel(trade)

    return closed

# ── Excel logging ─────────────────────────────────────────────────────────────

_HEADERS = [
    "Date", "Ticker", "Entry Price", "Exit Price", "Shares",
    "Position Size ($)", "P&L ($)", "Return (%)",
    "Hold Time (mins, total)",   # accumulates across all calendar days
    "Result", "Exit Reason",
]

def _create_excel_workbook():
    wb  = openpyxl.Workbook()

    # ── Trade Log sheet ────────────────────────────────────────────────────────
    ws  = wb.active
    ws.title = "Trade Log"

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    for col, hdr in enumerate(_HEADERS, 1):
        c = ws.cell(row=1, column=col, value=hdr)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    col_widths = [12, 8, 13, 13, 8, 18, 11, 13, 18, 14, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    # ── Summary sheet ──────────────────────────────────────────────────────────
    ss  = wb.create_sheet("Summary")
    ss["A1"] = "TJL Paper Trading — 15-Day Summary"
    ss["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ss.merge_cells("A1:B1")

    # Note: positions carry overnight — no EOD force-close
    ss["A2"] = ("Note: positions carry overnight/weekend until TP (+3%) or SL (-1.5%) is hit. "
                "Hold time includes all calendar minutes from entry to exit.")
    ss["A2"].font = Font(italic=True, size=9, color="595959")
    ss.merge_cells("A2:B2")

    metrics = [
        ("Total Trades",              "=COUNTA('Trade Log'!B2:B50000)"),
        ("Win Rate %",                "=IFERROR(COUNTIF('Trade Log'!J2:J50000,\"WIN\")"
                                      "/COUNTA('Trade Log'!J2:J50000)*100,0)"),
        ("Total P&L ($)",             "=SUM('Trade Log'!G2:G50000)"),
        ("Avg Return per Trade (%)",  "=IFERROR(AVERAGE('Trade Log'!H2:H50000),0)"),
        ("Avg Hold Time (mins)",      "=IFERROR(AVERAGE('Trade Log'!I2:I50000),0)"),
        ("Wins (TP hit)",             "=COUNTIF('Trade Log'!J2:J50000,\"WIN\")"),
        ("Losses (SL hit)",           "=COUNTIF('Trade Log'!J2:J50000,\"LOSS\")"),
    ]
    lbl_font  = Font(bold=True, size=11)
    val_fill  = PatternFill("solid", fgColor="EEF3FB")
    for row, (lbl, formula) in enumerate(metrics, 4):
        lc = ss.cell(row=row, column=1, value=lbl)
        lc.font = lbl_font
        vc = ss.cell(row=row, column=2, value=formula)
        vc.fill      = val_fill
        vc.alignment = Alignment(horizontal="right")
        if "%" in lbl:
            vc.number_format = "0.00\"%\""
        elif "$" in lbl or "P&L" in lbl:
            vc.number_format = '"$"#,##0.00'
        elif "mins" in lbl:
            vc.number_format = "0.0"

    ss.column_dimensions["A"].width = 30
    ss.column_dimensions["B"].width = 18

    wb.save(EXCEL_FILE)
    print(f"  [EXCEL] Workbook created -> {EXCEL_FILE}")
    return wb

def append_trade_to_excel(trade):
    if not OPENPYXL_OK:
        print(f"  [WARN] openpyxl unavailable — trade not logged to Excel: {trade}")
        return

    if not os.path.exists(EXCEL_FILE):
        _create_excel_workbook()

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Trade Log"]

    result_fills = {
        "WIN":  PatternFill("solid", fgColor="C6EFCE"),  # green
        "LOSS": PatternFill("solid", fgColor="FFC7CE"),  # red
    }
    row_fill = result_fills.get(trade["result"], PatternFill())
    next_row = ws.max_row + 1

    row_vals = [
        trade["date"],
        trade["symbol"],
        trade["entry_price"],
        trade["exit_price"],
        trade["shares"],
        trade["position_size"],
        trade["pnl"],
        trade["return_pct"],
        trade["hold_mins"],
        trade["result"],
        trade["exit_reason"],
    ]
    num_fmts = {
        3: "#,##0.0000", 4: "#,##0.0000",
        5: "0.00",
        6: '"$"#,##0.00',
        7: '"$"#,##0.00',
        8: '0.00"%"',
        9: "0.0",
    }
    for col, val in enumerate(row_vals, 1):
        c = ws.cell(row=next_row, column=col, value=val)
        c.fill = row_fill
        if col in num_fmts:
            c.number_format = num_fmts[col]

    wb.save(EXCEL_FILE)
    print(f"  [EXCEL] Row {next_row}: {trade['symbol']} {trade['result']} "
          f"P&L ${trade['pnl']:+.2f}")
